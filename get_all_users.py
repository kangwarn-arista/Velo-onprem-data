import json
import urllib3
import requests
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from requests.structures import CaseInsensitiveDict
import os

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings()

load_dotenv()

token = os.getenv("VCO_TOKEN")
vco_url = os.getenv("VCO_URL")

OUTPUT_XLSX = "vco_user_audit.xlsx"

ROLE_MAP = {
    "superuser": "Super User",
    "networkSuperSpecialist": "Super User",
    "networkSpecialist": "Standard Admin",
    "standard": "Standard Admin",
    "operator": "Standard Admin",
    "readonly": "Read Only",
    "businessSpecialist": "Standard Admin",
    "networkAdmin": "Standard Admin",
}


def api_call(method, params):
    headers = CaseInsensitiveDict()
    headers["Authorization"] = token
    headers["Content-Type"] = "application/x-www-form-urlencoded"

    data = {
        "id": 0,
        "jsonrpc": "2.0",
        "method": method,
        "params": params
    }

    try:
        resp = requests.post(vco_url, headers=headers, data=json.dumps(data), verify=False)
        return resp.json()
    except Exception as e:
        print(f"API Error: {e}")
        return {}


def get_enterprise_ids():
    parsed = api_call("network/getNetworkEnterprises", {"networkId": 1, "with": ["edges"]})
    return [
        {
            "id": item.get("id"),
            "name": item.get("name"),
            "logicalId": item.get("logicalId", "")
        }
        for item in parsed.get("result", [])
    ]


def get_enterprise_details(ent_id):
    return api_call("enterprise/getEnterprise", {
        "id": ent_id,
        "with": ["enterpriseProxy"]
    })


def get_enterprise_users(ent_id):
    parsed = api_call("enterprise/getEnterpriseUsers", {"enterpriseId": ent_id})
    return parsed.get("result", [])


def get_network_operator_users():
    parsed = api_call("network/getNetworkOperatorUsers", {"networkId": 1})
    return parsed.get("result", [])


def map_role(raw_role):
    if not raw_role:
        return "Unknown"
    return ROLE_MAP.get(raw_role, raw_role)


def safe_get(d, *keys):
    for key in keys:
        if isinstance(d, dict):
            d = d.get(key, "")
        else:
            return ""
    return d if d is not None else ""


# ── Styling helpers ──────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F3864")   # dark navy
ALT_FILL      = PatternFill("solid", start_color="DCE6F1")   # light blue
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROLE_COLORS = {
    "Super User":     "FF4444",
    "Standard Admin": "4472C4",
    "Read Only":      "70AD47",
    "Unknown":        "808080",
}


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, col_count, alternate=False):
    fill = ALT_FILL if alternate else WHITE_FILL
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = CELL_FONT
        cell.fill = fill
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_role_badge(ws, row, col, role):
    cell = ws.cell(row=row, column=col, value=role)
    color = ROLE_COLORS.get(role, "808080")
    cell.font = Font(name="Arial", bold=True, size=10, color=color)
    cell.alignment = CENTER


def write_section_title(ws, row, col_count, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=12, color="1F3864")
    cell.alignment = LEFT
    cell.fill = PatternFill("solid", start_color="EEF2F8")
    cell.border = THIN_BORDER


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_enterprise_sheet(wb, enterprise_rows):
    ws = wb.create_sheet("Enterprise Users")
    ws.freeze_panes = "A2"

    headers = [
        "Enterprise Name", "Partner Name",
        "Username", "Email", "First Name", "Last Name", "Role"
    ]
    widths = [30, 25, 25, 35, 18, 18, 18]

    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(enterprise_rows, start=2):
        ws.append([
            r["enterprise_name"], r["partner_name"],
            r["username"], r["email"], r["first_name"], r["last_name"],
            r["role"]
        ])
        style_data_row(ws, i, len(headers), alternate=(i % 2 == 0))
        add_role_badge(ws, i, 7, r["role"])

    set_col_widths(ws, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return ws


def build_partner_sheet(wb, partner_rows):
    ws = wb.create_sheet("Partner (Operator) Users")
    ws.freeze_panes = "A2"

    headers = [
        "Email", "First Name", "Last Name", "Role"
    ]
    widths = [35, 18, 18, 25]

    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(partner_rows, start=2):
        ws.append([
            r["email"], r["first_name"], r["last_name"], r["role"]
        ])
        style_data_row(ws, i, len(headers), alternate=(i % 2 == 0))
        add_role_badge(ws, i, 4, r["role"])

    set_col_widths(ws, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return ws


def build_summary_sheet(wb, enterprise_rows, partner_rows):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    def add_kv(label, value, bold_val=False):
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        v = ws.cell(row=r, column=2, value=value)
        v.font = BOLD_FONT if bold_val else CELL_FONT
        for c in [1, 2]:
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).alignment = LEFT

    write_section_title(ws, 1, 2, "VCO User Audit — Summary")

    ws.append([])
    write_section_title(ws, ws.max_row + 1, 2, "Partner (Operator) Users")
    add_kv("Total Partner Users", len(partner_rows))
    role_counts = {}
    for r in partner_rows:
        role_counts[r["role"]] = role_counts.get(r["role"], 0) + 1
    for role, cnt in sorted(role_counts.items()):
        add_kv(f"  {role}", cnt)

    ws.append([])
    write_section_title(ws, ws.max_row + 1, 2, "Enterprise Users")
    add_kv("Total Enterprises", len(set(r["enterprise_id"] for r in enterprise_rows)))
    add_kv("Total Enterprise Users", len(enterprise_rows))
    role_counts2 = {}
    for r in enterprise_rows:
        role_counts2[r["role"]] = role_counts2.get(r["role"], 0) + 1
    for role, cnt in sorted(role_counts2.items()):
        add_kv(f"  {role}", cnt)


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":

    if not token:
        print("ERROR: VCO_TOKEN not found in .env")
        exit(1)
    if not vco_url:
        print("ERROR: VCO_URL not found in .env")
        exit(1)

    # ── 1. Enterprises ──
    print("Fetching enterprises...")
    enterprises = get_enterprise_ids()
    print(f"  Found {len(enterprises)} enterprises")

    # ── 2. Partner name per enterprise ──
    print("Fetching partner names for each enterprise...")
    partner_map = {}   # enterprise_id -> partner_name
    for ent in enterprises:
        details = get_enterprise_details(ent["id"])
        partner_name = safe_get(details.get("result", {}), "enterpriseProxy", "name")
        partner_map[ent["id"]] = partner_name or ""

    # ── 3. Partner (operator) users ──
    print("Fetching network operator (partner) users...")
    operator_users = get_network_operator_users()
    print(f"  Found {len(operator_users)} operator users")

    partner_rows = []
    for u in operator_users:
        role_name = u.get("roleName", "")
        partner_rows.append({
            "username":   u.get("username", ""),
            "email":      u.get("email", ""),
            "first_name": u.get("firstName", ""),
            "last_name":  u.get("lastName", ""),
            "role":       role_name,
            "raw_role":   role_name,
        })

    # ── 4. Enterprise users ──
    enterprise_rows = []
    for ent in enterprises:
        print(f"  Fetching users for: {ent['name']} (id={ent['id']})")
        users = get_enterprise_users(ent["id"])
        print(f"    {len(users)} users found")
        for u in users:
            role_name = u.get("roleName", "")
            enterprise_rows.append({
                "enterprise_name": ent["name"],
                "partner_name":    partner_map.get(ent["id"], ""),
                "enterprise_id":   ent["id"],
                "username":        u.get("username", ""),
                "email":           u.get("email", ""),
                "first_name":      u.get("firstName", ""),
                "last_name":       u.get("lastName", ""),
                "role":            role_name,
                "raw_role":        role_name,
            })

    # ── 5. Write XLSX ──
    print("\nWriting Excel output...")
    wb = Workbook()

    build_summary_sheet(wb, enterprise_rows, partner_rows)
    build_enterprise_sheet(wb, enterprise_rows)
    build_partner_sheet(wb, partner_rows)

    wb.save(OUTPUT_XLSX)
    print(f"Done. Written to '{OUTPUT_XLSX}'")
    print(f"  Enterprise users : {len(enterprise_rows)}")
    print(f"  Partner users    : {len(partner_rows)}")
